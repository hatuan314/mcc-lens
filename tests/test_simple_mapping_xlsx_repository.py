"""Unit tests for SimpleMappingXlsxRepository."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.models.mapping_entry import MappingEntry, RankedMcc
from app.repositories.simple_mapping_xlsx_repository import SimpleMappingXlsxRepository


def _make_entry(vsic_code: str, vsic_title: str, mcc_code: str = "") -> MappingEntry:
    top_results = []
    if mcc_code:
        top_results = [
            RankedMcc(mcc_code=mcc_code, mcc_title="Test MCC", score=0.8, comment="ok")
        ]
    return MappingEntry(
        vsic_code=vsic_code, vsic_title=vsic_title, top_results=top_results
    )


class TestSimpleMappingXlsxRepository:
    def test_writes_header_row(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        out = tmp_path / "out.xlsx"
        repo.write([], out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(1, 1).value == "VSIC"
        assert ws.cell(1, 2).value == "MCC"
        assert ws.cell(1, 3).value == "Tên ngành"

    def test_writes_data_rows(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        entries = [
            _make_entry("0111", "Trồng lúa", "0111"),
            _make_entry("5411", "Bán lẻ thực phẩm", "5411"),
        ]
        out = tmp_path / "out.xlsx"
        repo.write(entries, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data
        assert ws.cell(2, 1).value == "0111"
        assert ws.cell(2, 2).value == "0111"
        assert ws.cell(2, 3).value == "Trồng lúa"

    def test_no_match_entry_has_empty_mcc(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        entries = [_make_entry("9999", "Không xác định")]
        out = tmp_path / "out.xlsx"
        repo.write(entries, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 2).value in ("", None)

    def test_creates_parent_directory(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        out = tmp_path / "subdir" / "out.xlsx"
        repo.write([], out)
        assert out.exists()

    def test_preserves_vietnamese_characters(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        title = "Hoạt động dịch vụ cá nhân khác chưa được phân vào đâu"
        entries = [_make_entry("9609", title, "7299")]
        out = tmp_path / "out.xlsx"
        repo.write(entries, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 3).value == title

    def test_top1_mcc_used_when_multiple_results(self, tmp_path: Path) -> None:
        repo = SimpleMappingXlsxRepository()
        entry = MappingEntry(
            vsic_code="0111",
            vsic_title="Trồng lúa",
            top_results=[
                RankedMcc(
                    mcc_code="0111", mcc_title="Farms", score=0.92, comment="best"
                ),
                RankedMcc(
                    mcc_code="5411", mcc_title="Grocery", score=0.61, comment="ok"
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        repo.write([entry], out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 2).value == "0111"

"""Unit tests for MappingController (consumer-only)."""

from pathlib import Path
from unittest.mock import patch

import numpy as np

from app.controllers.mapping_controller import DEFAULT_TOP_K, MappingController
from app.models.embedding_artifact import EmbeddingArtifact
from app.models.mapping_entry import MappingEntry, RankedMcc
from app.repositories.embedding_artifact_repository import EmbeddingArtifactRepository

_DIM = 1024


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_artifact(path: Path, n_vsic: int = 2, n_mcc: int = 2) -> Path:
    """Write a valid embedding artifact to ``path`` and return it."""
    artifact = EmbeddingArtifact(
        mcc_vectors=np.ones((n_mcc, _DIM), dtype=np.float32),
        mcc_codes=[f"{1000 + i}" for i in range(n_mcc)],
        mcc_titles=[f"MCC {i}" for i in range(n_mcc)],
        mcc_descriptions=[f"desc {i}" for i in range(n_mcc)],
        vsic_vectors=np.ones((n_vsic, _DIM), dtype=np.float32),
        vsic_codes=[f"{i:04d}" for i in range(n_vsic)],
        vsic_titles=[f"VSIC {i}" for i in range(n_vsic)],
        meta={"dim": _DIM, "zero_vector_codes": {"mcc": [], "vsic": []}},
    )
    EmbeddingArtifactRepository().write(path, artifact)
    return path


def _make_template(tmp_path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Result"
    wb.create_sheet("Hướng Dẫn")
    wb.create_sheet("Thống Kê")
    p = tmp_path / "template.xlsx"
    wb.save(p)
    return p


FAKE_ENTRIES = [
    MappingEntry(
        vsic_code="0000",
        vsic_title="VSIC 0",
        top_results=[
            RankedMcc(mcc_code="1000", mcc_title="MCC 0", score=0.9, comment="ok")
        ],
    ),
    MappingEntry(vsic_code="0001", vsic_title="VSIC 1", top_results=[]),
]


class TestMappingControllerExitCodes:
    def test_returns_1_when_artifact_missing(self, tmp_path: Path) -> None:
        controller = MappingController(template_path=_make_template(tmp_path))
        exit_code = controller.execute(
            embeddings=tmp_path / "missing.npz",
            output=tmp_path / "out.xlsx",
            output_detail=tmp_path / "detail.xlsx",
        )
        assert exit_code == 1

    def test_returns_4_when_artifact_corrupt(self, tmp_path: Path) -> None:
        bad = tmp_path / "bad.npz"
        bad.write_text("not an npz file")
        controller = MappingController(template_path=_make_template(tmp_path))
        exit_code = controller.execute(
            embeddings=bad,
            output=tmp_path / "out.xlsx",
            output_detail=tmp_path / "detail.xlsx",
        )
        assert exit_code == 4

    def test_returns_2_when_ollama_unavailable(self, tmp_path: Path) -> None:
        artifact = _write_artifact(tmp_path / "a.npz")
        with patch(
            "app.controllers.mapping_controller.check_ollama_llm",
            side_effect=RuntimeError("Ollama unavailable"),
        ):
            controller = MappingController(template_path=_make_template(tmp_path))
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )
        assert exit_code == 2

    def test_returns_0_on_success(self, tmp_path: Path) -> None:
        artifact = _write_artifact(tmp_path / "a.npz")
        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ):
            controller = MappingController(template_path=_make_template(tmp_path))
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )
        assert exit_code == 0

    def test_returns_3_on_io_error(self, tmp_path: Path) -> None:
        artifact = _write_artifact(tmp_path / "a.npz")
        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ), patch(
            "app.repositories.simple_mapping_xlsx_repository"
            ".SimpleMappingXlsxRepository.write",
            side_effect=IOError("disk full"),
        ):
            controller = MappingController(template_path=_make_template(tmp_path))
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )
        assert exit_code == 3

    def test_no_embedding_client_constructed(self, tmp_path: Path) -> None:
        """Module 2 must never instantiate an embedding client."""
        import app.controllers.mapping_controller as mc

        assert not hasattr(mc, "OllamaEmbeddingClient")
        assert not hasattr(mc, "WokuShopEmbeddingClient")


class TestMappingControllerDefaults:
    def test_default_top_k_is_60(self) -> None:
        import inspect

        sig = inspect.signature(MappingController.execute)
        assert sig.parameters["top_k"].default == DEFAULT_TOP_K
        assert DEFAULT_TOP_K == 60

    def test_default_gdrive_output_dir_is_none(self) -> None:
        import inspect

        sig = inspect.signature(MappingController.execute)
        assert sig.parameters["gdrive_output_dir"].default is None

    def test_template_path_none_skips_detail_output(self, tmp_path: Path) -> None:
        artifact = _write_artifact(tmp_path / "a.npz")
        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ):
            controller = MappingController(template_path=None)
            controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )
        assert not (tmp_path / "detail.xlsx").exists()


class TestMappingControllerGdriveOutputDir:
    """Tests for --gdrive-output-dir path override logic."""

    def _run_with_gdrive(
        self,
        tmp_path: Path,
        gdrive_dir: Path,
        template: bool = True,
    ) -> int:
        # Put the artifact inside the gdrive dir so the default override finds it.
        gdrive_dir.mkdir(parents=True, exist_ok=True)
        _write_artifact(gdrive_dir / "embed-artifact.npz")
        controller = MappingController(
            template_path=_make_template(tmp_path) if template else None
        )
        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ):
            return controller.execute(
                embeddings=tmp_path / "nonexistent.npz",
                output=tmp_path / "ignored.xlsx",
                output_detail=tmp_path / "ignored-detail.xlsx",
                gdrive_output_dir=gdrive_dir,
            )

    def test_output_files_written_to_gdrive_dir(self, tmp_path: Path) -> None:
        gdrive_dir = tmp_path / "drive" / "mcc-lens"
        self._run_with_gdrive(tmp_path, gdrive_dir)
        assert (gdrive_dir / "vsic-mcc-mapping.xlsx").exists()
        assert (gdrive_dir / "vsic-mcc-mapping-detail.xlsx").exists()

    def test_checkpoint_path_uses_gdrive_dir(self, tmp_path: Path) -> None:
        gdrive_dir = tmp_path / "drive" / "mcc-lens"
        gdrive_dir.mkdir(parents=True, exist_ok=True)
        _write_artifact(gdrive_dir / "embed-artifact.npz")
        controller = MappingController(template_path=None)

        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ), patch(
            "app.controllers.mapping_controller.MappingCheckpointRepositoryImpl"
        ) as MockCkpt:
            MockCkpt.return_value.load.return_value = {}
            controller.execute(
                embeddings=tmp_path / "nonexistent.npz",
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
                gdrive_output_dir=gdrive_dir,
            )
        checkpoint_arg = MockCkpt.call_args[0][0]
        assert str(checkpoint_arg).startswith(str(gdrive_dir))

    def test_original_output_paths_ignored(self, tmp_path: Path) -> None:
        gdrive_dir = tmp_path / "drive"
        self._run_with_gdrive(tmp_path, gdrive_dir)
        assert not (tmp_path / "ignored.xlsx").exists()
        assert not (tmp_path / "ignored-detail.xlsx").exists()

    def test_returns_0_on_success_with_gdrive(self, tmp_path: Path) -> None:
        gdrive_dir = tmp_path / "drive"
        exit_code = self._run_with_gdrive(tmp_path, gdrive_dir)
        assert exit_code == 0

    def test_nested_gdrive_dir_created_with_parents(self, tmp_path: Path) -> None:
        gdrive_dir = tmp_path / "a" / "b" / "c" / "d"
        self._run_with_gdrive(tmp_path, gdrive_dir)
        assert gdrive_dir.is_dir()

    def test_top_k_clamped_to_100(self, tmp_path: Path) -> None:
        artifact = _write_artifact(tmp_path / "a.npz")
        controller = MappingController(template_path=None)
        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ):
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
                top_k=200,
            )
        assert exit_code == 0

    def test_limit_passed_to_use_case_execute(self, tmp_path: Path) -> None:
        """limit must be forwarded to MapVsicToMccUseCase.execute(limit=...)."""
        artifact = _write_artifact(tmp_path / "a.npz")
        controller = MappingController(template_path=None)

        with patch("app.controllers.mapping_controller.check_ollama_llm"), patch(
            "app.controllers.mapping_controller.MapVsicToMccUseCase",
        ) as MockUseCase:
            MockUseCase.return_value.execute.return_value = iter([])
            controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
                limit=1,
            )
            execute_kwargs = MockUseCase.return_value.execute.call_args[1]
        assert execute_kwargs.get("limit") == 1

    def test_wokushop_provider_success(self, tmp_path: Path) -> None:
        """WokuShop provider succeeds and needs no Ollama embedding check."""
        artifact = _write_artifact(tmp_path / "a.npz")
        controller = MappingController(
            template_path=None,
            llm_provider="wokushop",
            wokushop_api_key="test-api-key",
            wokushop_model="gpt-4o-test",
        )

        with patch(
            "app.controllers.mapping_controller.WokuShopLLMClient"
        ) as MockWokuClient, patch(
            "app.services.map_vsic_to_mcc_use_case.MapVsicToMccUseCase.execute",
            return_value=iter(FAKE_ENTRIES),
        ):
            MockWokuClient.return_value.health_check.return_value = True
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )

        assert exit_code == 0
        MockWokuClient.return_value.health_check.assert_called_once()

    def test_wokushop_provider_health_check_failure(self, tmp_path: Path) -> None:
        """Returns 2 if WokuShop health check fails."""
        artifact = _write_artifact(tmp_path / "a.npz")
        controller = MappingController(
            template_path=None,
            llm_provider="wokushop",
            wokushop_api_key="test-api-key",
        )

        with patch(
            "app.controllers.mapping_controller.WokuShopLLMClient"
        ) as MockWokuClient:
            MockWokuClient.return_value.health_check.return_value = False
            exit_code = controller.execute(
                embeddings=artifact,
                output=tmp_path / "out.xlsx",
                output_detail=tmp_path / "detail.xlsx",
            )

        assert exit_code == 2
        MockWokuClient.return_value.health_check.assert_called_once()

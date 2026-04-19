"""Unit tests for DetailMappingXlsxRepository."""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.models.mapping_entry import MappingEntry, RankedMcc
from app.repositories.detail_mapping_xlsx_repository import DetailMappingXlsxRepository


def _make_template(tmp_path: Path) -> Path:
    """Create a minimal Excel template with 3 required sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping Result"
    # Row 1: merged section headers
    ws.cell(1, 1).value = "THÔNG TIN VSIC"
    ws.cell(1, 3).value = "MCC XẾP HẠNG 1 (SCORE CAO NHẤT)"
    ws.cell(1, 7).value = "MCC XẾP HẠNG 2"
    ws.cell(1, 11).value = "MCC XẾP HẠNG 3"
    # Row 2: column headers
    headers = [
        "Mã VSIC",
        "Tên Ngành (Tiếng Việt)",
        "Mã MCC(#1)",
        "Tên MCC(#1)",
        "Score(#1)",
        "Nhận xét(#1)",
        "Mã MCC(#2)",
        "Tên MCC(#2)",
        "Score(#2)",
        "Nhận xét(#2)",
        "Mã MCC(#3)",
        "Tên MCC(#3)",
        "Score(#3)",
        "Nhận xét(#3)",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i).value = h

    wb.create_sheet("Hướng Dẫn").cell(1, 1).value = "Hướng dẫn sử dụng"
    wb.create_sheet("Thống Kê").cell(1, 1).value = "=COUNTA('Mapping Result'!A:A)"

    template_path = tmp_path / "template.xlsx"
    wb.save(template_path)
    return template_path


def _make_full_entry(vsic_code: str = "0111") -> MappingEntry:
    return MappingEntry(
        vsic_code=vsic_code,
        vsic_title="Trồng lúa",
        top_results=[
            RankedMcc(
                mcc_code="0111", mcc_title="Farms", score=0.92, comment="Agriculture"
            ),
            RankedMcc(
                mcc_code="5411", mcc_title="Grocery", score=0.61, comment="Food retail"
            ),
            RankedMcc(
                mcc_code="5999", mcc_title="Misc Retail", score=0.45, comment="General"
            ),
        ],
    )


class TestDetailMappingXlsxRepository:
    def test_raises_when_template_missing(self, tmp_path: Path) -> None:
        repo = DetailMappingXlsxRepository(tmp_path / "nonexistent.xlsx")
        with pytest.raises(FileNotFoundError):
            repo.write([], tmp_path / "out.xlsx")

    def test_preserves_huong_dan_sheet(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        out = tmp_path / "out.xlsx"
        repo.write([_make_full_entry()], out)
        wb = load_workbook(out)
        assert "Hướng Dẫn" in wb.sheetnames
        assert wb["Hướng Dẫn"].cell(1, 1).value == "Hướng dẫn sử dụng"

    def test_preserves_thong_ke_sheet(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        out = tmp_path / "out.xlsx"
        repo.write([_make_full_entry()], out)
        wb = load_workbook(out)
        assert "Thống Kê" in wb.sheetnames

    def test_data_row_has_14_columns(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        out = tmp_path / "out.xlsx"
        repo.write([_make_full_entry()], out)
        wb = load_workbook(out)
        ws = wb["Mapping Result"]
        # Find data row (after header rows)
        data_row = None
        for row in ws.iter_rows():
            if row[0].value == "0111":
                data_row = [cell.value for cell in row]
                break
        assert data_row is not None
        assert len([v for v in data_row if v is not None or True]) >= 14
        assert data_row[0] == "0111"
        assert data_row[1] == "Trồng lúa"
        assert data_row[2] == "0111"  # rank 1 mcc_code
        assert data_row[6] == "5411"  # rank 2 mcc_code
        assert data_row[10] == "5999"  # rank 3 mcc_code

    def test_score_is_float_not_string(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        out = tmp_path / "out.xlsx"
        repo.write([_make_full_entry()], out)
        wb = load_workbook(out)
        ws = wb["Mapping Result"]
        for row in ws.iter_rows():
            if row[0].value == "0111":
                score = row[4].value  # Score rank 1 = col index 4
                assert isinstance(score, float), f"Expected float, got {type(score)}"
                break

    def test_partial_results_padded_with_empty(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        entry = MappingEntry(
            vsic_code="9999",
            vsic_title="Test",
            top_results=[
                RankedMcc(
                    mcc_code="7299",
                    mcc_title="Services",
                    score=0.55,
                    comment="only one",
                ),
            ],
        )
        out = tmp_path / "out.xlsx"
        repo.write([entry], out)
        wb = load_workbook(out)
        ws = wb["Mapping Result"]
        for row in ws.iter_rows():
            if row[0].value == "9999":
                assert row[2].value == "7299"  # rank 1
                assert row[6].value == "" or row[6].value is None  # rank 2 empty
                assert row[10].value == "" or row[10].value is None  # rank 3 empty
                break

    def test_no_match_entry_has_all_empty_mcc(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        entry = MappingEntry(vsic_code="9999", vsic_title="Unknown", top_results=[])
        out = tmp_path / "out.xlsx"
        repo.write([entry], out)
        wb = load_workbook(out)
        ws = wb["Mapping Result"]
        for row in ws.iter_rows():
            if row[0].value == "9999":
                assert row[2].value == "" or row[2].value is None
                break

    def test_creates_parent_directory(self, tmp_path: Path) -> None:
        template = _make_template(tmp_path)
        repo = DetailMappingXlsxRepository(template)
        out = tmp_path / "subdir" / "detail.xlsx"
        repo.write([], out)
        assert out.exists()

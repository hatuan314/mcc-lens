"""
Unit tests for VsicExcelRepository.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.repositories.vsic_excel_repository import VsicExcelRepository


@pytest.fixture
def repo() -> VsicExcelRepository:
    return VsicExcelRepository()


@pytest.fixture
def sample_excel(tmp_path: Path) -> Path:
    """Create a sample Excel file for testing."""
    excel_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active

    ws.append(["Mã ngành nghề ", "Tên ngành"])
    ws.append([1110, "Trồng lúa"])
    ws.append([1120, "Trồng ngô"])
    ws.append([1130, "Trồng cây lấy củ"])

    wb.save(excel_path)
    return excel_path


class TestVsicExcelRepository:
    def test_read_rows_returns_dict_list(
        self, repo: VsicExcelRepository, sample_excel: Path
    ) -> None:
        rows = repo.read_rows(sample_excel)

        assert isinstance(rows, list)
        assert len(rows) == 3
        assert all(isinstance(row, dict) for row in rows)

    def test_read_rows_headers_correct(
        self, repo: VsicExcelRepository, sample_excel: Path
    ) -> None:
        rows = repo.read_rows(sample_excel)

        assert "Mã ngành nghề " in rows[0]
        assert "Tên ngành" in rows[0]

    def test_read_rows_data_correct(
        self, repo: VsicExcelRepository, sample_excel: Path
    ) -> None:
        rows = repo.read_rows(sample_excel)

        assert rows[0]["Mã ngành nghề "] == 1110
        assert rows[0]["Tên ngành"] == "Trồng lúa"
        assert rows[1]["Mã ngành nghề "] == 1120
        assert rows[1]["Tên ngành"] == "Trồng ngô"

    def test_file_not_found(self, repo: VsicExcelRepository, tmp_path: Path) -> None:
        non_existent = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            repo.read_rows(non_existent)

    def test_empty_file(self, repo: VsicExcelRepository, tmp_path: Path) -> None:
        excel_path = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(excel_path)

        rows = repo.read_rows(excel_path)
        assert rows == []

    def test_skip_empty_rows(self, repo: VsicExcelRepository, tmp_path: Path) -> None:
        excel_path = tmp_path / "empty_rows.xlsx"
        wb = Workbook()
        ws = wb.active

        ws.append(["Mã ngành nghề ", "Tên ngành"])
        ws.append([1110, "Trồng lúa"])
        ws.append([None, None])
        ws.append([1120, "Trồng ngô"])

        wb.save(excel_path)

        rows = repo.read_rows(excel_path)
        assert len(rows) == 2

"""
VSIC 2025 Excel Repository implementation.
"""

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from loguru import logger


class Vsic2025ExcelRepository:
    """
    Repository for reading VSIC 2025 data from Excel file.

    Expected columns: Cấp 1, Cấp 2, Cấp 3, Cấp 4, Cấp 5, Tên ngành
    """

    EXPECTED_HEADERS = ["Cấp 1", "Cấp 2", "Cấp 3", "Cấp 4", "Cấp 5", "Tên ngành"]

    def read_rows(self, input_path: Path) -> List[Dict[str, Any]]:
        """
        Read rows from VSIC 2025 Excel file.

        Args:
            input_path: Path to the Excel file.

        Returns:
            List of row dictionaries with expected column headers as keys.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If headers don't match expected format.
        """
        if not input_path.exists():
            raise FileNotFoundError(f"File not found: {input_path}")

        workbook = load_workbook(input_path, data_only=True)
        sheet = workbook.active

        rows = []
        headers = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = self._validate_headers(row)
                logger.info(f"Validated headers: {headers}")
                continue

            # Skip empty rows (no code in any level column)
            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    row_dict[headers[col_idx]] = cell
                else:
                    row_dict[f"col_{col_idx}"] = cell

            rows.append(row_dict)

        workbook.close()
        logger.info(f"Read {len(rows)} data rows from {input_path}")
        return rows

    def _validate_headers(self, header_row: tuple) -> List[str]:
        """
        Validate that headers match expected format.

        Args:
            header_row: First row containing column headers.

        Returns:
            List of header names.

        Raises:
            ValueError: If headers don't match expected format.
        """
        headers = [
            str(cell).strip() if cell is not None else f"col_{i}"
            for i, cell in enumerate(header_row)
        ]

        # Check for required headers
        required = ["Cấp 4", "Cấp 5", "Tên ngành"]
        missing = [h for h in required if h not in headers]

        if missing:
            raise ValueError(
                f"Missing required headers: {missing}. "
                f"Found headers: {headers}"
            )

        return headers

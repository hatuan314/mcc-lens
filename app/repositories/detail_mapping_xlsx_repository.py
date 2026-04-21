"""Repository for writing detailed mapping Excel file using template."""

from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from app.models.mapping_entry import MappingEntry


class DetailMappingXlsxRepository:
    """Repository for writing detailed VSIC-MCC mapping Excel file."""

    def __init__(self, template_path: Path) -> None:
        """
        Initialize repository with template path.

        Args:
            template_path: Path to Excel template file.
        """
        self.template_path = template_path

    def write(self, entries: list[MappingEntry], output_path: Path) -> None:
        """
        Write mapping entries to detailed Excel file using template.

        Loads template, fills "Mapping Result" sheet with 14 columns,
        preserves "Hướng Dẫn" and "Thống Kê" sheets unchanged.

        Args:
            entries: List of MappingEntry objects.
            output_path: Path to output Excel file.
        """
        logger.info(
            f"Writing detailed mapping file to {output_path} with "
            f"{len(entries)} entries"
        )

        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")

        # Load template
        wb = load_workbook(self.template_path)

        # Get or create Mapping Result sheet
        if "Mapping Result" in wb.sheetnames:
            ws = wb["Mapping Result"]
            # Clear existing data (keep first 3 rows as header area)
            if ws.max_row >= 4:
                ws.delete_rows(4, ws.max_row)
        else:
            ws = wb.create_sheet("Mapping Result")
            # Ensure data starts at row 4 by adding 2 empty rows then header
            ws.append([])  # Row 1
            ws.append([])  # Row 2
            # Header at Row 3
            header = [
                "Mã VSIC",
                "Tên Ngành (Tiếng Việt)",
                # Rank 1
                "Mã MCC (Rank 1)",
                "Tên MCC (Rank 1)",
                "Score (Rank 1)",
                "Nhận xét (Rank 1)",
                # Rank 2
                "Mã MCC (Rank 2)",
                "Tên MCC (Rank 2)",
                "Score (Rank 2)",
                "Nhận xét (Rank 2)",
                # Rank 3
                "Mã MCC (Rank 3)",
                "Tên MCC (Rank 3)",
                "Score (Rank 3)",
                "Nhận xét (Rank 3)",
            ]
            ws.append(header)

        # Data rows
        for entry in entries:
            row: list = [
                entry.vsic_code,
                entry.vsic_title,
            ]

            # Pad top_results to exactly 3 ranks
            ranked = list(entry.top_results[:3])
            for rank in ranked:
                row.extend([rank.mcc_code, rank.mcc_title, rank.score, rank.comment])
            for _ in range(3 - len(ranked)):
                row.extend(["", "", "", ""])

            ws.append(row)

        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(output_path)
        logger.info(f"Successfully wrote detailed mapping file to {output_path}")

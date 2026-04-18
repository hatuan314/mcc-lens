"""
VSIC Excel Repository implementation.
"""

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


class VsicExcelRepository:
    """
    Repository for reading VSIC data from Excel file.
    """

    def read_rows(self, input_path: Path) -> List[Dict[str, Any]]:
        """
        Read rows from VSIC Excel file.

        Args:
            input_path: Path to the Excel file.

        Returns:
            List of row dictionaries with column headers as keys.
        """
        if not input_path.exists():
            raise FileNotFoundError(f"File not found: {input_path}")

        workbook = load_workbook(input_path, data_only=True)
        sheet = workbook.active

        rows = []
        headers = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [cell if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
                continue

            if not any(cell is not None for cell in row):
                continue

            row_dict = {}
            for col_idx, cell in enumerate(row):
                key = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                row_dict[key] = cell

            rows.append(row_dict)

        workbook.close()
        return rows

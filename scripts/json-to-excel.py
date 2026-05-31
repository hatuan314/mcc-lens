import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_PATH  = "output/vsicvn-mcc-mapping-detail.json"
VSIC_PATH   = "output/vsic-vn.json"
MCC_PATH    = "output/mcc-visa.json"
OUTPUT_PATH = "output/vsicvn-mcc-mapping.xlsx"

# ── Colors ──────────────────────────────────────────────────────────────────
HEADER_BG   = "1F4E79"   # dark blue
HEADER_FG   = "FFFFFF"
RANK1_BG    = "E2EFDA"   # light green
RANK2_BG    = "FFF2CC"   # light yellow
RANK3_BG    = "FCE4D6"   # light orange
ALT_BG      = "F5F5F5"
WHITE_BG    = "FFFFFF"

THIN_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

RANK_COLORS = {1: RANK1_BG, 2: RANK2_BG, 3: RANK3_BG}

HEADERS = [
    "VSIC Code", "VSIC Name",
    "Rank", "MCC Code", "MCC Name", "Score", "Comment"
]
COL_WIDTHS = [12, 32, 8, 12, 42, 9, 60]


def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def style_cell(cell, bold=False, bg=None, fg="000000", wrap=False,
               h_align="left", v_align="top", border=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = make_fill(bg)
    cell.alignment = Alignment(
        horizontal=h_align, vertical=v_align, wrap_text=wrap
    )
    if border:
        cell.border = THIN_BORDER


def build_workbook(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data.get("sheet_name", "Mapping Result")

    # ── Metadata rows ────────────────────────────────────────────────────────
    meta_font = Font(name="Arial", size=10, italic=True, color="595959")
    for label, value in [
        ("Mô tả:", data.get("description", "")),
        ("Nguồn:",  data.get("source_file", "")),
        ("Tổng số bản ghi:", str(len(data["mappings"]))),
    ]:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10, bold=True, color="595959")
        c = ws.cell(row=row, column=2, value=value)
        c.font = meta_font
        c.alignment = Alignment(horizontal="left", wrap_text=True)

    ws.append([])  # blank spacer

    # ── Header row ───────────────────────────────────────────────────────────
    header_row = ws.max_row + 1
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        style_cell(c, bold=True, bg=HEADER_BG, fg=HEADER_FG,
                   h_align="center", v_align="center")
    ws.row_dimensions[header_row].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    mappings = data["mappings"]
    for mapping in mappings:
        vsic_code = mapping["vsic_code"]
        vsic_name = mapping["vsic_name"]
        candidates = mapping["mcc_candidates"]

        first_data_row = ws.max_row + 1

        for cand in candidates:
            rank       = cand["rank"]
            mcc_code   = cand["mcc_code"]
            mcc_name   = cand["mcc_name"]
            score      = cand["score"]
            comment    = cand["comment"]

            row_num = ws.max_row + 1
            row_bg  = RANK_COLORS.get(rank, WHITE_BG)

            values = [vsic_code, vsic_name, rank, mcc_code, mcc_name, score, comment]
            for col_idx, val in enumerate(values, start=1):
                c = ws.cell(row=row_num, column=col_idx, value=val)
                h_align = "center" if col_idx in (1, 3, 4, 6) else "left"
                style_cell(c, bg=row_bg, wrap=(col_idx == 7), h_align=h_align)
            ws.row_dimensions[row_num].height = 60

        # Merge VSIC code and VSIC name columns across 3 candidate rows
        last_data_row = ws.max_row
        if last_data_row > first_data_row:
            for col in (1, 2):
                ws.merge_cells(
                    start_row=first_data_row, start_column=col,
                    end_row=last_data_row,    end_column=col
                )
                c = ws.cell(row=first_data_row, column=col)
                style_cell(
                    c, bg=RANK_COLORS[1],
                    h_align="center", v_align="center",
                    bold=(col == 1)
                )

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Auto-filter on header ─────────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row}"
    )

    return wb


def normalize_mapping(mapping: dict) -> dict:
    """Normalize legacy format (vsic_title/suggested_mcc) to standard format."""
    if "vsic_name" in mapping:
        return mapping
    candidates = []
    for cand in mapping.get("suggested_mcc", []):
        candidates.append({
            "rank": cand.get("rank", 0),
            "mcc_code": cand.get("mcc_code", ""),
            "mcc_name": cand.get("mcc_description", ""),
            "score": 0.0,
            "comment": cand.get("reason", ""),
        })
    return {
        "vsic_code": mapping.get("vsic_code", ""),
        "vsic_name": mapping.get("vsic_title", ""),
        "mcc_candidates": candidates,
    }


def load_vsic_lookup() -> dict:
    with open(VSIC_PATH, encoding="utf-8") as f:
        vsic_data = json.load(f)
    return {item["code"]: item["title"] for item in vsic_data["vsic_list"]}


def load_mcc_lookup() -> dict:
    with open(MCC_PATH, encoding="utf-8") as f:
        mcc_data = json.load(f)
    return {item["mcc"]: item["title"] for item in mcc_data["mcc_list"]}


def main():
    with open(INPUT_PATH, encoding="utf-8") as f:
        data = json.load(f)

    vsic_lookup = load_vsic_lookup()
    mcc_lookup = load_mcc_lookup()
    data["mappings"] = [normalize_mapping(m) for m in data["mappings"]]
    for m in data["mappings"]:
        title = vsic_lookup.get(m["vsic_code"])
        if title:
            m["vsic_name"] = title
        for cand in m.get("mcc_candidates", []):
            mcc_title = mcc_lookup.get(cand["mcc_code"])
            if mcc_title:
                cand["mcc_name"] = mcc_title

    wb = build_workbook(data)
    wb.save(OUTPUT_PATH)
    print(f"Saved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()